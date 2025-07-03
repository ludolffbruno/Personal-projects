# ------------------------------------------------------------
# IMPORTAÇÕES NECESSÁRIAS
# ------------------------------------------------------------
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import pyperclip

# Função para validar entrada no "Custo Unitário"
def validar_custo(P):
    if P == "" or P.replace(",", "").replace(".", "").isdigit():
        return True
    return False

# Variáveis globais
dados = []
dados_planilha = []  # Nova variável para armazenar os dados completos da planilha

# Função para carregar os dados da planilha
def carregar_dados():
    global dados, dados_planilha
    try:
        workbook = openpyxl.load_workbook("CALCULO_ST_DIFAL_COMPLETA.xlsx", data_only=True)
        sheet = workbook["COMPLETA"]
        dados.clear()
        dados_planilha.clear()
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            ncm = str(row[0])
            produto = row[1]
            st_difal = row[2]
            aquisicao = row[3]  # Coluna D
            origem = row[4]     # Coluna E
            if ncm and produto:
                dados.append({"NCM": ncm, "Produto": produto, "ST / DIFAL": st_difal, "Aquisição": aquisicao, "Origem": origem})
                dados_planilha.append(row)
    except FileNotFoundError:
        resumo_var.set("Arquivo CALCULO_ST.xlsx não encontrado!")
    except Exception as e:
        resumo_var.set(f"Erro ao carregar dados: {str(e)}")

# Função para filtrar os dados
def filtrar_dados(event=None):
    global resultados_tree
    consulta = busca_var.get().lower().replace(".", "")
    resultados_tree.delete(*resultados_tree.get_children())
    for item in dados:
        ncm = item["NCM"].replace(".", "").lower()
        produto = item["Produto"].lower()
        if consulta in ncm or consulta in produto:
            resultados_tree.insert("", "end", values=(item["NCM"], item["Produto"]))

# Função para exibir detalhes do item selecionado e atualizar campos dependentes
def selecionar_item(event):
    global resultados_tree, dados_planilha
    item_selecionado = resultados_tree.focus()
    if item_selecionado:
        valores = resultados_tree.item(item_selecionado, "values")
        ncm_var.set(valores[0])  # Preencher o NCM
        resumo_var.set(f"NCM {valores[0]} | Produto: {valores[1]}")
        
        # Usar os dados já carregados em dados_planilha
        ncm_input = valores[0].replace(".", "")
        for row in dados_planilha:
            ncm_planilha = str(row[0]).replace(".", "")
            if ncm_planilha == ncm_input:
                st_difal = row[2] or "N/A"  # Coluna C (índice 2)
                st_difal_var.set(str(st_difal))
                atualizar_campos_dependentes()
                break
        else:
            st_difal_var.set("N/A")
            atualizar_campos_dependentes()

# Função para atualizar os campos dependentes com base em ST / DIFAL
def atualizar_campos_dependentes():
    if st_difal_var.get() == "C/ST":
        st_inclusa_checkbutton.config(state="normal")
        if st_inclusa_var.get():
            origem_icms_combobox.config(state="disabled")
        else:
            origem_icms_combobox.config(state="normal")  # Desbloqueado para seleção do usuário
    else:  # S/ST ou outro valor
        st_inclusa_var.set(False)  # Desmarcar o checkbox
        st_inclusa_checkbutton.config(state="disabled")
        origem_icms_combobox.config(state="normal")  # Habilitar Origem ICMS para S/ST

# Função para atualizar Origem ICMS com base em ST Inclusa na compra?
def atualizar_origem_icms(*args):
    if st_difal_var.get() == "C/ST":
        if st_inclusa_var.get():  # Se o checkbox está marcado
            origem_icms_var.set("0%")  # Valor padrão quando desabilitado
            origem_icms_combobox.config(state="disabled")
        else:
            origem_icms_combobox.config(state="normal")  # Desbloqueado para seleção do usuário
    else:
        origem_icms_combobox.config(state="normal")  # Sempre habilitado para S/ST

# Função para calcular o custo final
def calcular_custo():
    global dados_planilha
    try:
        ncm_input = ncm_var.get().replace(".", "")
        custo_unitario_str = custo_var.get().replace(",", ".")
        custo_unitario = float(custo_unitario_str)

        linha_ncm = None
        for row in dados_planilha:
            ncm_planilha = str(row[0]).replace(".", "")
            if ncm_planilha == ncm_input:
                linha_ncm = row
                break

        if not linha_ncm:
            resumo_var.set("NCM não encontrado na planilha.")
            return

        # Extrair dados da planilha
        aquisicao = linha_ncm[3]  # Coluna D: INTERNA ou EXTERNA
        origem = linha_ncm[4]     # Coluna E: NACIONAL ou IMPORTADA
        st_inclusa = linha_ncm[5] # Coluna F: SIM, NÃO ou ICMS-ST
        origem_importada = linha_ncm[6] or 0  # Coluna G (4%)
        origem_nacional = linha_ncm[7] or 0   # Coluna H (12%)
        origem_rj = linha_ncm[8] or 0         # Coluna I (22%)

        # Determinar origem_icms_value para C/ST com ST não inclusa
        if st_difal_var.get() == "C/ST" and not st_inclusa_var.get():
            origem_icms_percent = origem_icms_var.get()
            if origem_icms_percent == "4%":
                origem_icms_value = origem_importada / 100
            elif origem_icms_percent == "12%":
                origem_icms_value = origem_nacional / 100
            elif origem_icms_percent == "22%":
                origem_icms_value = origem_rj / 100
            else:
                origem_icms_value = 0
        else:
            # Para C/ST com ST inclusa ou S/ST, determinamos automaticamente
            if origem == "IMPORTADA":
                origem_icms_value = origem_importada / 100
            elif origem == "NACIONAL":
                origem_icms_value = origem_nacional / 100
            else:
                origem_icms_value = 0

        if st_difal_var.get() == "C/ST":
            if st_inclusa_var.get():  # ST já recolhido
                custo_st_interes_estadual = 0
                custo_st_interna = 0
            else:  # ST não inclusa
                custo_st_interes_estadual = custo_unitario * origem_icms_value
                custo_st_interna = 0  # Removido o cálculo de ST interna
        else:  # S/ST
            origem_icms_percent = origem_icms_var.get()
            if origem_icms_percent == "4%":
                origem_icms_value = origem_importada / 100
            elif origem_icms_percent == "12%":
                origem_icms_value = origem_nacional / 100
            elif origem_icms_percent == "22%":
                origem_icms_value = origem_rj / 100
            else:
                origem_icms_value = 0
            custo_st_interes_estadual = custo_unitario * origem_icms_value
            custo_st_interna = 0

        custo_final = custo_unitario + custo_st_interes_estadual + custo_st_interna
        custo_final_formatado = f"{custo_final:.2f}".replace('.', ',')
        custo_calculado_var.set(custo_final_formatado)
        
    except ValueError:
        resumo_var.set("Erro nos dados de entrada. Verifique os valores!")
    except Exception as e:
        resumo_var.set(f"Erro: {str(e)}")

# Função para copiar o custo final
def copiar_custo():
    try:
        root.clipboard_clear()
        root.clipboard_append(custo_calculado_var.get())
        root.update()
        resumo_var.set("Custo final copiado para a área de transferência!")
    except Exception as e:
        resumo_var.set(f"Erro ao copiar: {str(e)}")

# Função para limpar os campos
def limpar_campos():
    ncm_var.set("")
    st_difal_var.set("N/A")
    st_inclusa_var.set(False)  # Desmarcar o checkbox
    origem_icms_var.set("0%")
    custo_var.set("")
    custo_calculado_var.set("")
    resumo_var.set("")
    resultados_tree.selection_remove(resultados_tree.selection())
    st_inclusa_checkbutton.config(state="disabled")
    origem_icms_combobox.config(state="normal")  # Habilitar Origem ICMS por padrão
    resultados_tree.delete(*resultados_tree.get_children())
    for item in dados:
        resultados_tree.insert("", "end", values=(item["NCM"], item["Produto"]))

# Função para exibir o tutorial
def mostrar_tutorial():
    
    tutorial = (
        "Tutorial: Como Usar o Programa de Cálculo ST/DIFAL para Compras\n\n"
        "1. Abra o Programa\n"
        "- Inicie o programa. A janela abrirá maximizada automaticamente.\n\n"
        "2. Busque o NCM ou Produto\n"
        "- No campo 'Buscar (NCM ou Produto)', digite o NCM ou nome do produto (ex.: '7605' ou 'fios').\n"
        "- Clique no item desejado na tabela que aparece abaixo.\n\n"
        "3. Verifique os Dados Automáticos\n"
        "- 'NCM Selecionado' e 'ST / DIFAL' (C/ST ou S/ST) serão preenchidos automaticamente.\n"
        "- Se 'ST / DIFAL' for 'C/ST', o campo 'ST Inclusa na compra?' será ativado.\n\n"
        "4. Configure 'ST Inclusa na compra?' (se aplicável)\n\n"
        "- Para 'C/ST': Marque a caixa se o ST interestadual já foi recolhido.\n\n"
        "  - Se marcada, 'Origem ICMS' será desativado e o custo final será o custo unitário.\n\n"
        "  - Se não marcada, escolha a 'Origem ICMS' (0%, 4%, 12% ou 22%) para calcular o ST interestadual.\n\n\n"
        "5. Escolha 'Origem ICMS' (se aplicável)\n"
        "- Para 'S/ST' ou 'C/ST' com 'ST Inclusa' desmarcada: Selecione o percentual de 'Origem ICMS' (0%, 4%, 12% ou 22%).\n\n"
        "6. Insira o Custo Unitário\n"
        "- Digite o valor do produto no campo 'Custo Unitário' (ex.: '1200,50'). Use vírgula para decimais.\n\n"
        "7. Calcule o Custo Final\n"
        "- Clique em 'Calcular' ou pressione 'Enter' no campo 'Custo Unitário'.\n"
        "- O 'Custo Final' será exibido, refletindo o ST interestadual conforme a configuração (sem ST interna adicional).\n\n"
        "8. Copie o Resultado (Opcional)\n"
        "- Clique em 'Copiar Custo' para copiar o custo final para a área de transferência.\n\n"
        "9. Limpe os Campos (Opcional)\n"
        "- Clique em 'Limpar' para reiniciar todos os campos e a tabela.\n\n"
        "Dica\n"
        "- Certifique-se de que a planilha 'CALCULO_ST_DIFAL_COMPLETA.xlsx' está na mesma pasta do programa."
    )
    tk.messagebox.showinfo("Tutorial", tutorial)

# ------------------------------------------------------------
# VARIÁVEIS GLOBAIS E INTERFACE GRÁFICA
# ------------------------------------------------------------
root = tk.Tk()
root.title("CÁLCULO ST/DIFAL")
root.option_add("*Font", ("Arial", 16))  # Reduz a fonte global para 16

# Maximizar a janela automaticamente
root.state('zoomed')  # Abre maximizada no Windows
root.minsize(900, 700)  # Tamanho mínimo

busca_var = tk.StringVar()
ncm_var = tk.StringVar()
custo_var = tk.StringVar()
resumo_var = tk.StringVar()
custo_calculado_var = tk.StringVar(value="")
st_difal_var = tk.StringVar(value="N/A")
st_inclusa_var = tk.BooleanVar(value=False)  # Booleano para o checkbox
origem_icms_var = tk.StringVar(value="0%")

# Validação de custo
vcmd = (root.register(validar_custo), '%P')

frame = ttk.Frame(root, padding="10")
frame.grid(sticky="NSEW")

# Configurar expansão do frame
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

# Título
ttk.Label(frame, text="CÁLCULO ST/DIFAL", font=("Arial", 18, "bold"), foreground="blue").grid(column=0, row=0, columnspan=4, pady=5, sticky="EW")

# Botão Tutorial no canto superior direito
ttk.Button(frame, text="Tutorial", command=mostrar_tutorial).grid(column=3, row=0, sticky="NE", padx=5, pady=5)

# Buscar
ttk.Label(frame, text="Buscar (NCM ou Produto):").grid(column=0, row=1, sticky="W", padx=5, pady=2)
entry_busca = ttk.Entry(frame, textvariable=busca_var)
entry_busca.grid(column=1, row=1, sticky="EW", columnspan=3, padx=5, pady=2)
entry_busca.bind("<KeyRelease>", filtrar_dados)

style = ttk.Style()
style.theme_use("clam")  # Tema mais moderno
style.configure("Treeview", font=("Arial", 14), rowheight=30)
style.configure("Treeview.Heading", font=("Arial", 14, "bold"))
style.configure("TButton", font=("Arial", 14))
style.configure("Large.TCheckbutton", font=("Arial", 10)) #botão flag

# Criar Treeview com scrollbar
colunas = ("NCM", "Produto")
resultados_tree = ttk.Treeview(frame, columns=colunas, show="headings", height=12)  # Primeiro cria o Treeview
resultados_tree.heading("NCM", text="NCM")
resultados_tree.heading("Produto", text="Produto")
resultados_tree.column("NCM", width=260, minwidth=80, stretch=False)  # Reduzido para 80 pixels
resultados_tree.column("Produto", width=550, minwidth=550)
resultados_tree.grid(column=0, row=2, columnspan=4, padx=5, pady=5, sticky="NSEW")

scrollbar = ttk.Scrollbar(frame, orient="vertical", command=resultados_tree.yview)  # Depois cria o Scrollbar
scrollbar.grid(column=4, row=2, sticky="NS", padx=5, pady=5)
resultados_tree.configure(yscrollcommand=scrollbar.set)  # Vincula o yscrollcommand após criar o Treeview

resultados_tree.bind("<<TreeviewSelect>>", selecionar_item)

# NCM Selecionado
ttk.Label(frame, text="NCM Selecionado:").grid(column=0, row=3, sticky="W", padx=5, pady=2)
ttk.Entry(frame, textvariable=ncm_var, state="readonly").grid(column=1, row=3, sticky="EW", columnspan=3, padx=5, pady=2)

# ST / DIFAL
ttk.Label(frame, text="ST / DIFAL:").grid(column=0, row=4, sticky="W", padx=5, pady=2)
ttk.Entry(frame, textvariable=st_difal_var, state="readonly").grid(column=1, row=4, sticky="EW", columnspan=3, padx=5, pady=2)

# ST Inclusa na compra? (checkbox)
ttk.Label(frame, text="ST Inclusa na compra?:").grid(column=0, row=5, sticky="W", padx=5, pady=2)
st_inclusa_checkbutton = ttk.Checkbutton(frame, variable=st_inclusa_var, state="disabled", style="Large.TCheckbutton")
st_inclusa_checkbutton.grid(column=1, row=5, sticky="W", padx=5, pady=2)
st_inclusa_var.trace("w", atualizar_origem_icms)

# Origem ICMS (atualizado para 4%, 12%, 22%)
ttk.Label(frame, text="Origem ICMS:").grid(column=0, row=6, sticky="W", padx=5, pady=2)
origem_icms_combobox = ttk.Combobox(frame, textvariable=origem_icms_var, values=["0%", "4%", "12%", "22%"], state="normal", font=("Arial", 14))
origem_icms_combobox.grid(column=1, row=6, sticky="EW", columnspan=3, padx=5, pady=2)

# Custo Unitário
ttk.Label(frame, text="Custo Unitário:").grid(column=0, row=7, sticky="W", padx=5, pady=2)
custo_entry = ttk.Entry(frame, textvariable=custo_var, font=("Arial", 14), validate="key", validatecommand=vcmd)
custo_entry.grid(column=1, row=7, sticky="EW", columnspan=3, padx=5, pady=2)
custo_entry.bind("<Return>", lambda event: calcular_custo())  # Adiciona suporte à tecla Enter

# Botões
ttk.Button(frame, text="Calcular", command=calcular_custo).grid(column=0, row=8, sticky="W", padx=5, pady=2)
ttk.Button(frame, text="Copiar Custo", command=copiar_custo).grid(column=1, row=8, sticky="EW", padx=5, pady=2)
ttk.Button(frame, text="Limpar", command=limpar_campos).grid(column=2, row=8, sticky="E", padx=5, pady=2)

# Selecionados
ttk.Label(frame, text="Selecionados:", foreground="blue", font=("Arial", 14)).grid(column=0, row=9, columnspan=4, sticky="W", padx=5, pady=2)
ttk.Label(frame, textvariable=resumo_var, font=("Arial", 14), wraplength=700).grid(column=0, row=10, columnspan=4, sticky="NSEW", padx=5, pady=2)

# Custo Final
ttk.Label(frame, text="Custo Final:", foreground="green", font=("Arial", 14)).grid(column=0, row=11, sticky="W", padx=5, pady=2)
ttk.Label(frame, textvariable=custo_calculado_var, foreground="green", font=("Arial", 14)).grid(column=1, row=11, sticky="W", columnspan=3, padx=5, pady=2)

carregar_dados()
filtrar_dados()

# Configurar expansão do frame e widgets
frame.columnconfigure(1, weight=1)  # Expande a segunda coluna (widgets com fundo branco)
frame.columnconfigure(2, weight=1)  # Expande a terceira coluna (mais espaço para os widgets)
frame.columnconfigure(3, weight=1)  # Expande a quarta coluna (para alinhamento)
frame.rowconfigure(2, weight=1)    # Faz o Treeview expandir verticalmente
frame.rowconfigure(10, weight=1)   # Faz o label de "Selecionados" expandir verticalmente

root.mainloop()
